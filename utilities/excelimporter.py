from openpyxl import load_workbook

from models import Category, Question


def migrate_from_excel_2_db(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    row = 3
    cat = Category.create(name=ws["A1"].value)
    while True:
        if ws['A' + str(row)].value == None:
            break
        if ws['B' + str(row)].value == "input":
            Question.create(value=ws['C' + str(row)].value,
                            correct_answer=ws['D' + str(row)].value,
                            type="input", category_id=cat.id)
        elif ws['B' + str(row)].value == "choose":
            Question.create(type="choose", category_id=cat.id,
                            correct_answer=ws['D' + str(row)].value,
                            value=ws['C' + str(row)].value,
                            wrong_answers=ws['E'+str(row)].value.split(", "),
                            answers_amount=ws['F'+str(row)].value)
        row += 1
